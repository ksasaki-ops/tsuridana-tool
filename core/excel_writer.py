"""
Excel 発注書生成モジュール（openpyxl）
"""
import os
import shutil

from openpyxl import load_workbook

from core.text_builder import OrderSpec
from utils.path_utils import asset, downloads_dir


def generate_excel(spec: OrderSpec, output_path: str | None = None) -> str:
    """
    テンプレートをコピーして発注書Excelを生成し、パスを返す。
    output_path が None の場合は Downloads フォルダに保存。
    """
    template_path = asset("家具発注_template.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"Excelテンプレートが見つかりません: {template_path}\n"
            "assets/ フォルダに 家具発注_template.xlsx を配置してください。"
        )

    if output_path is None:
        filename = f"家具発注_{spec.customer}様.xlsx"
        output_path = os.path.join(downloads_dir(), filename)

    # テンプレートをコピーしてから編集（元ファイルを汚さない）
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    # C6: お客様名｜物件名
    cell_customer = ws["C6"]
    cell_customer.value = f"{spec.customer}様｜{spec.property_name}"

    # L3: 発注No.
    ws["L3"].value = f"発注No.：{spec.hassou_no}"

    # L4: 発注日
    ws["L4"].value = f"発注日：{spec.hassou_date}"

    wb.save(output_path)
    wb.close()

    return output_path
